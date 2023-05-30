import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

def count_representatives(filename):
    """
    Function reads JSON data from a file and generates an Excel table that shows the count of representatives in different files.

    Author:
    Mia Jurdana
    """
    unique_filenames = set()
    with open(filename) as file:
        json_data = json.load(file)
    representative_info = {}
    for obj in json_data:
        filename = obj["fileName"].split("_")[0]
        unique_filenames.add(filename)
        representatives = obj["representatives"]

        for representative in representatives:
            if representative in representative_info:
                representative_info[representative]["count"] += 1
                representative_info[representative]["filenames"].append(filename)
            else:
                representative_info[representative] = {
                    "count": 1,
                    "filenames": [filename]
                }

    all_representatives = sorted(representative_info.keys())
    all_filenames = sorted(unique_filenames)

    table_data = []
    for representative in all_representatives:
        row = []
        row.append(representative)
        for filename in all_filenames:
            if filename in representative_info[representative]["filenames"]:
                row.append("+")
            else:
                row.append("")
        row.append(representative_info[representative]["count"])
        table_data.append(row)

    df = pd.DataFrame(table_data, columns=["Representative"] + all_filenames + ["Count"])

    wb = Workbook()
    ws = wb.active

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    max_text_length = max([len(str(cell.value)) for cell in ws["A"]])
    ws.column_dimensions["A"].width = max_text_length

    for cell in ws["A"]:
        cell.alignment = Alignment(wrap_text=True)

    wb.save("./results/representative_table.xlsx")

    print("Excel table generated successfully!")
    
    # for representative, info in representative_info.items():
    #     count = info["count"]
    #     filenames = "\n".join(info["filenames"])
    #     print(
    #         f"Representative '{representative}' exists in {count} files: \n {filenames}")
    # print(unique_filenames)